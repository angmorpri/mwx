# 2026/04/20
"""
import_.py - Import wallet data from an Excel workbook.

Counterpart of export.py. Reads a .xlsx file previously produced by
Wallet.export() (or hand-crafted following the same schema) and
applies the changes to a Wallet instance in place.

Key design decisions
--------------------
- MWID is the anchor: rows with a MWID update the existing entity;
  rows without MWID create new entities (mwid=-1, which the model
  auto-assigns elsewhere).
- Missing entities (present in wallet, absent from Excel) are NOT
  deleted by default. Pass `delete_missing=True` to enable.
- Errors fail fast with a message identifying sheet, row and column.
- The Tipo column is read but ignored on updates (immutable field).
- The Detalles column restores newlines from the ' // ' separator.
- Capital validation is opt-in via `validate=True`, and compares
  per-account + total against the canonical values stored in the
  hidden __meta__ sheet (tolerance: 0.01 €).

"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import TYPE_CHECKING, Any

from openpyxl import load_workbook

from mwx.etl.xexport import compute_capital
from mwx.model import Account, Category, Counterpart, Entry
from mwx.util import Money

if TYPE_CHECKING:
    from mwx.wallet import Wallet


# Constants (must stay in sync with export.py)

_META_SHEET_NAME = "__meta__"
_META_SCHEMA_SUPPORTED = {"mwx-export/1"}
_CAPITAL_TOLERANCE = 0.01  # €, at-cent precision

_TYPE_FROM_LABEL = {
    "Gasto": -1,
    "Transferencia": 0,
    "Ingreso": +1,
}


# Exceptions


class ImportError_(Exception):
    """Base exception for import failures. Renamed to avoid builtin clash."""


class CellError(ImportError_):
    """Error tied to a specific cell: carries sheet/row/column/value info."""

    def __init__(
        self, sheet: str, row: int, column: str, value: Any, reason: str
    ) -> None:
        self.sheet = sheet
        self.row = row
        self.column = column
        self.value = value
        self.reason = reason
        super().__init__(f"[{sheet}!{column}{row}] {reason} (value: {value!r})")


class ValidationError(ImportError_):
    """Capital validation failed: wallet totals don't match __meta__."""


# Public data classes


@dataclass
class ImportDiff:
    """Summary of what an import would do / has done.

    Contains the entities to add, modify and delete, grouped by type.
    Returned by import_() in dry_run mode, and also after a real
    import (for logging).

    """

    accounts_add: list[Account] = field(default_factory=list)
    accounts_modify: list[tuple[Account, Account]] = field(default_factory=list)
    accounts_delete: list[Account] = field(default_factory=list)

    categories_add: list[Category] = field(default_factory=list)
    categories_modify: list[tuple[Category, Category]] = field(default_factory=list)
    categories_delete: list[Category] = field(default_factory=list)

    entries_add: list[Entry] = field(default_factory=list)
    entries_modify: list[tuple[Entry, Entry]] = field(default_factory=list)
    entries_delete: list[Entry] = field(default_factory=list)

    counterparts_new: list[Counterpart] = field(default_factory=list)

    capital_check: dict[str, Any] | None = None  # set if validate=True

    def summary(self) -> str:
        """Produce a human-readable multi-line summary."""
        lines = []
        lines.append(
            f"Cuentas:    +{len(self.accounts_add)}  "
            f"~{len(self.accounts_modify)}  -{len(self.accounts_delete)}"
        )
        lines.append(
            f"Categorías: +{len(self.categories_add)}  "
            f"~{len(self.categories_modify)}  -{len(self.categories_delete)}"
        )
        lines.append(
            f"Movimientos: +{len(self.entries_add)}  "
            f"~{len(self.entries_modify)}  -{len(self.entries_delete)}"
        )
        if self.counterparts_new:
            lines.append(f"Contrapartes nuevas: {len(self.counterparts_new)}")
        if self.capital_check:
            status = self.capital_check.get("status", "not checked")
            lines.append(f"Validación capital: {status}")
            if self.capital_check.get("discrepancies"):
                for d in self.capital_check["discrepancies"]:
                    lines.append(f"  ⚠ {d}")
        return "\n".join(lines)


# Reading helpers


def _read_sheet_rows(ws, expected_headers: list[str]) -> list[dict[str, Any]]:
    """Read a sheet as a list of dicts keyed by header name.

    Row 1 is the header row. Rows below are data. Empty rows (all
    cells None) are skipped. Each dict also has `__row__` with the
    Excel row number (for error messages).

    """
    if ws.max_row < 1:
        return []

    # Read header
    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [h.strip() if isinstance(h, str) else h for h in header_cells]

    # Sanity check: every expected header must be present
    missing = [h for h in expected_headers if h not in headers]
    if missing:
        raise ImportError_(
            f"Sheet '{ws.title}' missing required columns: {missing}. "
            f"Found: {headers}"
        )

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(c is None for c in row):
            continue
        data = {h: v for h, v in zip(headers, row) if h is not None}
        data["__row__"] = row_idx
        rows.append(data)
    return rows


def _read_meta(ws) -> dict[str, Any]:
    """Read the __meta__ sheet. Layout: two tables separated by blank row.

    Top table: key/value metadata.
    Bottom table: per-account capital.

    Returns a dict with keys 'metadata' and 'per_account_capital'.

    """
    metadata: dict[str, Any] = {}
    per_account: dict[str, float] = {}

    in_meta = True  # we start in the top table (skipping its header row)
    saw_per_account_header = False

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # Header rows of either table
        if row_idx == 1:  # "Meta" / "Valor"
            continue
        # Blank row = separator
        if all(c is None for c in row):
            in_meta = False
            continue
        key, val = row[0], row[1] if len(row) > 1 else None
        if in_meta:
            if key is not None:
                metadata[str(key)] = val
        else:
            if not saw_per_account_header:
                # This is the "Cuenta" / "Capital" header
                saw_per_account_header = True
                continue
            if key is not None:
                per_account[str(key)] = float(val) if val is not None else 0.0

    return {"metadata": metadata, "per_account_capital": per_account}


# Field coercion helpers


def _as_str(value: Any, sheet: str, row: int, col: str, *, default: str = "") -> str:
    if value is None:
        return default
    return str(value).strip()


def _as_mwid(value: Any, sheet: str, row: int, col: str) -> int:
    """None/empty → -1 (new entity). Otherwise int."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return -1
    try:
        return int(value)
    except (ValueError, TypeError):
        raise CellError(sheet, row, col, value, "MWID debe ser entero o vacío")


def _as_bool(
    value: Any, sheet: str, row: int, col: str, *, default: bool = False
) -> bool:
    """None/empty → default. Booleans pass through. Strings parsed flexibly."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        v = value.strip().lower()
        if v in ("true", "verdadero", "sí", "si", "1", "yes"):
            return True
        if v in ("false", "falso", "no", "0"):
            return False
    raise CellError(sheet, row, col, value, "Valor booleano no reconocido")


def _as_int(
    value: Any, sheet: str, row: int, col: str, *, default: int | None = None
) -> int:
    if value is None or (isinstance(value, str) and not value.strip()):
        if default is not None:
            return default
        raise CellError(sheet, row, col, value, "Entero requerido, está vacío")
    try:
        return int(value)
    except (ValueError, TypeError):
        raise CellError(sheet, row, col, value, "No es un entero válido")


def _as_money(value: Any, sheet: str, row: int, col: str) -> Money:
    if value is None or (isinstance(value, str) and not value.strip()):
        raise CellError(sheet, row, col, value, "Importe requerido, está vacío")
    try:
        return Money(value)
    except Exception as e:
        raise CellError(sheet, row, col, value, f"Importe inválido: {e}")


def _as_date(value: Any, sheet: str, row: int, col: str):
    from datetime import date, datetime

    if value is None:
        raise CellError(sheet, row, col, value, "Fecha requerida, está vacía")
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, str):
        v = value.strip()
        for fmt in ("%Y-%m-%d", "%Y%m%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(v, fmt)
            except ValueError:
                continue
    raise CellError(sheet, row, col, value, "Fecha no reconocida")


def _as_color(
    value: Any, sheet: str, row: int, col: str, *, default: str = "#000000"
) -> str:
    if value is None or (isinstance(value, str) and not value.strip()):
        return default
    v = str(value).strip()
    if not v.startswith("#"):
        v = "#" + v
    if len(v) != 7:
        raise CellError(sheet, row, col, value, "Color debe ser formato '#RRGGBB'")
    return v.upper()


def _restore_newlines(value: Any) -> str:
    """Reverse of export's ' // ' flattening."""
    if value is None:
        return ""
    s = str(value)
    # Accept ' // ', '// ', ' //', '//' — all map to a single newline
    for sep in (" // ", "// ", " //", "//"):
        if sep in s:
            return s.replace(sep, "\n")
    return s


# Builders (rows → model objects)


def _build_account(row: dict, sheet: str) -> Account:
    r = row["__row__"]
    mwid = _as_mwid(row.get("MWID"), sheet, r, "A")
    name = _as_str(row.get("Nombre"), sheet, r, "B")
    if not name:
        raise CellError(sheet, r, "B", row.get("Nombre"), "Nombre requerido")
    order = _as_int(row.get("Orden"), sheet, r, "C", default=-1)
    color = _as_color(row.get("Color"), sheet, r, "D")
    is_visible = _as_bool(row.get("Visible"), sheet, r, "E", default=True)
    is_legacy = _as_bool(row.get("Legacy"), sheet, r, "F", default=False)
    try:
        return Account(
            mwid=mwid,
            name=name,
            order=order,
            color=color,
            is_visible=is_visible,
            is_legacy=is_legacy,
        )
    except ValueError as e:
        raise CellError(sheet, r, "?", None, f"Constructor Account falló: {e}")


def _build_category(row: dict, sheet: str) -> Category:
    r = row["__row__"]
    mwid = _as_mwid(row.get("MWID"), sheet, r, "A")
    code = _as_str(row.get("Código"), sheet, r, "B")
    if not code:
        raise CellError(sheet, r, "B", row.get("Código"), "Código requerido")
    name = _as_str(row.get("Nombre"), sheet, r, "C")
    if not name:
        raise CellError(sheet, r, "C", row.get("Nombre"), "Nombre requerido")
    # Type is read-only from the user's perspective: for new categories
    # we still need to know it, so we read it from the label
    tipo_label = _as_str(row.get("Tipo"), sheet, r, "D")
    if tipo_label not in _TYPE_FROM_LABEL:
        raise CellError(
            sheet,
            r,
            "D",
            tipo_label,
            f"Tipo debe ser uno de {list(_TYPE_FROM_LABEL)}",
        )
    cat_type = _TYPE_FROM_LABEL[tipo_label]
    icon_id = _as_int(row.get("Icono"), sheet, r, "E", default=0)
    color = _as_color(row.get("Color"), sheet, r, "F")
    is_legacy = _as_bool(row.get("Legacy"), sheet, r, "G", default=False)
    try:
        return Category(
            mwid=mwid,
            repr_name=f"{code}. {name}",
            cat_type=cat_type,
            icon_id=icon_id,
            color=color,
            is_legacy=is_legacy,
        )
    except ValueError as e:
        raise CellError(sheet, r, "?", None, f"Constructor Category falló: {e}")


def _resolve_account(
    ref: str, accounts_by_name: dict[str, Account], sheet: str, row: int, col: str
) -> Account:
    """Resolve an account reference (repr_name like '@Banco') to an Account."""
    if ref in accounts_by_name:
        return accounts_by_name[ref]
    raise CellError(
        sheet,
        row,
        col,
        ref,
        f"Cuenta '{ref}' no existe. Cuentas disponibles: "
        f"{sorted(accounts_by_name)}",
    )


def _resolve_category(
    ref: str, cats_by_key: dict[str, Category], sheet: str, row: int, col: str
) -> Category:
    """Resolve a category reference by code OR repr_name."""
    if ref in cats_by_key:
        return cats_by_key[ref]
    raise CellError(
        sheet,
        row,
        col,
        ref,
        f"Categoría '{ref}' no existe. Usa el código (e.g. 'E01') o el "
        f"nombre completo (e.g. 'E01. Comida').",
    )


def _build_entry(
    row: dict,
    sheet: str,
    accounts_by_name: dict[str, Account],
    cats_by_key: dict[str, Category],
    counterparts_by_name: dict[str, Counterpart],
) -> tuple[Entry, list[Counterpart]]:
    """Build an Entry from a row. Returns the entry and any newly
    created counterparts (so the caller can track them).
    """
    r = row["__row__"]
    mwid = _as_mwid(row.get("MWID"), sheet, r, "A")
    date = _as_date(row.get("Fecha"), sheet, r, "B")

    tipo_label = _as_str(row.get("Tipo"), sheet, r, "C")
    if tipo_label not in _TYPE_FROM_LABEL:
        raise CellError(
            sheet,
            r,
            "C",
            tipo_label,
            f"Tipo debe ser uno de {list(_TYPE_FROM_LABEL)}",
        )
    ent_type = _TYPE_FROM_LABEL[tipo_label]

    amount = _as_money(row.get("Importe"), sheet, r, "D")
    src_ref = _as_str(row.get("Origen"), sheet, r, "E")
    tgt_ref = _as_str(row.get("Destino"), sheet, r, "F")
    cat_ref = _as_str(row.get("Categoría"), sheet, r, "G")

    if not src_ref:
        raise CellError(sheet, r, "E", None, "Origen requerido")
    if not tgt_ref:
        raise CellError(sheet, r, "F", None, "Destino requerido")
    if not cat_ref:
        raise CellError(sheet, r, "G", None, "Categoría requerida")

    category = _resolve_category(cat_ref, cats_by_key, sheet, r, "G")

    # Source/target depend on entry type:
    #   +1 income:   source=Counterpart, target=Account
    #   -1 expense:  source=Account,     target=Counterpart
    #    0 transfer: source=Account,     target=Account
    new_counterparts: list[Counterpart] = []

    def _get_or_create_counterpart(name: str) -> Counterpart:
        if name in counterparts_by_name:
            return counterparts_by_name[name]
        cp = Counterpart(name=name)
        counterparts_by_name[name] = cp
        new_counterparts.append(cp)
        return cp

    if ent_type == +1:  # income
        source = _get_or_create_counterpart(src_ref)
        target = _resolve_account(tgt_ref, accounts_by_name, sheet, r, "F")
    elif ent_type == -1:  # expense
        source = _resolve_account(src_ref, accounts_by_name, sheet, r, "E")
        target = _get_or_create_counterpart(tgt_ref)
    else:  # transfer
        source = _resolve_account(src_ref, accounts_by_name, sheet, r, "E")
        target = _resolve_account(tgt_ref, accounts_by_name, sheet, r, "F")

    item = _as_str(row.get("Concepto"), sheet, r, "H")  # "" → model defaults
    details_raw = row.get("Detalles")
    details = _restore_newlines(details_raw)
    is_bill = _as_bool(row.get("Es factura"), sheet, r, "J", default=False)

    try:
        entry = Entry(
            mwid=mwid,
            amount=amount,
            date=date,
            ent_type=ent_type,
            source=source,
            target=target,
            category=category,
            item=item,
            details=details,
            is_bill=is_bill,
        )
    except ValueError as e:
        raise CellError(sheet, r, "?", None, f"Constructor Entry falló: {e}")

    return entry, new_counterparts


# Diff computation


def _diff_entities(
    old: list, new: list, *, ignore_mwid_zero: bool = False
) -> tuple[list, list[tuple], list]:
    """Compute (adds, modifies, deletes) between two entity lists by MWID.

    - `adds`: entities in `new` whose mwid == -1 (new entities).
    - `modifies`: pairs (old_entity, new_entity) with matching mwid > 0
       where the entity's to_dict() differs.
    - `deletes`: entities in `old` whose mwid is not present in `new`.

    `ignore_mwid_zero` is used for Counterparts (their mwid is always 0).

    """
    if ignore_mwid_zero:
        # For counterparts we don't diff here; they're handled separately.
        return [], [], []

    old_by_mwid = {e.mwid: e for e in old if e.mwid >= 0}
    new_with_mwid = [e for e in new if e.mwid >= 0]
    new_by_mwid = {e.mwid: e for e in new_with_mwid}

    adds = [e for e in new if e.mwid == -1]
    modifies = []
    for mwid, new_e in new_by_mwid.items():
        if mwid in old_by_mwid:
            old_e = old_by_mwid[mwid]
            if old_e.to_dict() != new_e.to_dict():
                modifies.append((old_e, new_e))
    deletes = [e for e in old if e.mwid >= 0 and e.mwid not in new_by_mwid]
    return adds, modifies, deletes


# Public API


def import_(
    wallet: "Wallet",
    path: str | Path,
    *,
    validate: bool = False,
    delete_missing: bool = False,
    dry_run: bool = False,
) -> ImportDiff:
    """Import wallet data from an Excel workbook.

    Reads an .xlsx file produced by Wallet.export() (or hand-crafted
    with the same schema) and applies the changes to `wallet` in
    place. By default, entities missing from the Excel are preserved
    (safer default); pass `delete_missing=True` to remove them.

    Parameters
    ----------
    wallet:
        The Wallet to modify in place.
    path:
        Path to the .xlsx file.
    validate:
        If True, check that per-account and total capital match the
        canonical values stored in the __meta__ sheet (tolerance 0.01).
        Raises ValidationError on mismatch.
    delete_missing:
        If True, entities present in `wallet` but missing from the
        Excel are deleted. If False (default), they are preserved and
        a warning is included in the returned diff.
    dry_run:
        If True, compute the diff and return it without modifying
        `wallet`. Useful for previewing changes.

    Returns
    -------
    ImportDiff
        A summary of the changes (applied, or that would be applied
        in dry_run mode). Call `diff.summary()` for a human-readable
        report.

    Raises
    ------
    CellError
        If a specific cell has an invalid value. Aborts on first error.
    ValidationError
        If `validate=True` and capital totals don't match.
    ImportError_
        For structural problems (missing sheet, bad schema, etc.).

    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Import source not found: {path}")

    wb = load_workbook(path, data_only=True)

    # Schema check (soft: missing __meta__ is fine unless validate=True)
    meta_info = None
    if _META_SHEET_NAME in wb.sheetnames:
        meta_info = _read_meta(wb[_META_SHEET_NAME])
        schema = meta_info["metadata"].get("schema")
        if schema is not None and schema not in _META_SCHEMA_SUPPORTED:
            raise ImportError_(
                f"Unsupported meta schema '{schema}'. "
                f"Supported: {_META_SCHEMA_SUPPORTED}"
            )
    elif validate:
        raise ImportError_(
            "Cannot validate: '__meta__' sheet missing. "
            "This file wasn't produced by Wallet.export()."
        )

    # Required visible sheets
    for required in ("Cuentas", "Categorías", "Movimientos"):
        if required not in wb.sheetnames:
            raise ImportError_(f"Required sheet '{required}' missing")

    # --- Read rows from each sheet ---
    account_rows = _read_sheet_rows(
        wb["Cuentas"],
        ["MWID", "Nombre", "Orden", "Color", "Visible", "Legacy"],
    )
    category_rows = _read_sheet_rows(
        wb["Categorías"],
        ["MWID", "Código", "Nombre", "Tipo", "Icono", "Color", "Legacy"],
    )
    entry_rows = _read_sheet_rows(
        wb["Movimientos"],
        [
            "MWID",
            "Fecha",
            "Tipo",
            "Importe",
            "Origen",
            "Destino",
            "Categoría",
            "Concepto",
            "Detalles",
            "Es factura",
        ],
    )

    # --- Build new entities (fail-fast) ---
    new_accounts = [_build_account(r, "Cuentas") for r in account_rows]
    new_categories = [_build_category(r, "Categorías") for r in category_rows]

    # Lookup tables for entry resolution. Include BOTH old and new names
    # so that entries referencing a just-renamed account by its old name
    # still work. The NEW names win on collision.
    accounts_by_name: dict[str, Account] = {}
    for a in wallet.accounts:
        accounts_by_name[a.repr_name] = a
    for a in new_accounts:
        # For existing mwids, find the corresponding old account to preserve
        # identity. Otherwise, use the new one directly.
        if a.mwid >= 0:
            existing = next((x for x in wallet.accounts if x.mwid == a.mwid), None)
            if existing is not None:
                accounts_by_name[a.repr_name] = existing
            else:
                accounts_by_name[a.repr_name] = a
        else:
            accounts_by_name[a.repr_name] = a

    cats_by_key: dict[str, Category] = {}
    for c in wallet.categories:
        cats_by_key[c.code] = c
        cats_by_key[c.repr_name] = c
    for c in new_categories:
        if c.mwid >= 0:
            existing = next((x for x in wallet.categories if x.mwid == c.mwid), None)
            target = existing if existing is not None else c
        else:
            target = c
        cats_by_key[c.code] = target
        cats_by_key[c.repr_name] = target

    # Counterparts: keyed by name. Start with existing, add new on the fly.
    counterparts_by_name: dict[str, Counterpart] = {
        cp.name: cp for cp in wallet.counterparts
    }
    all_new_counterparts: list[Counterpart] = []

    new_entries: list[Entry] = []
    for r in entry_rows:
        entry, new_cps = _build_entry(
            r,
            "Movimientos",
            accounts_by_name,
            cats_by_key,
            counterparts_by_name,
        )
        new_entries.append(entry)
        all_new_counterparts.extend(new_cps)

    # --- Compute diff ---
    diff = ImportDiff()
    diff.accounts_add, diff.accounts_modify, diff.accounts_delete = _diff_entities(
        wallet.accounts, new_accounts
    )
    diff.categories_add, diff.categories_modify, diff.categories_delete = (
        _diff_entities(wallet.categories, new_categories)
    )
    diff.entries_add, diff.entries_modify, diff.entries_delete = _diff_entities(
        wallet.entries, new_entries
    )
    diff.counterparts_new = all_new_counterparts

    # --- Optional capital validation ---
    if validate:
        # Simulate applying the diff onto a throwaway copy and compute capital
        simulated = _simulate_apply(wallet, diff, delete_missing=delete_missing)
        got_total, got_per_acc = compute_capital(simulated)

        expected_total = meta_info["metadata"].get("total_capital")
        expected_per_acc = meta_info["per_account_capital"]

        discrepancies = []
        if expected_total is not None:
            if abs(got_total - float(expected_total)) > _CAPITAL_TOLERANCE:
                discrepancies.append(
                    f"Capital total: esperado {expected_total:.2f}, "
                    f"calculado {got_total:.2f} "
                    f"(diff {got_total - float(expected_total):+.2f})"
                )

        for name, expected in expected_per_acc.items():
            got = got_per_acc.get(name)
            if got is None:
                discrepancies.append(
                    f"Cuenta '{name}' esperada en meta pero ausente tras import"
                )
                continue
            if abs(got - expected) > _CAPITAL_TOLERANCE:
                discrepancies.append(
                    f"Cuenta {name}: esperado {expected:.2f}, "
                    f"calculado {got:.2f} (diff {got - expected:+.2f})"
                )

        diff.capital_check = {
            "status": "OK" if not discrepancies else "FAIL",
            "expected_total": expected_total,
            "got_total": got_total,
            "discrepancies": discrepancies,
        }

        if discrepancies:
            raise ValidationError(
                "Capital validation failed:\n  " + "\n  ".join(discrepancies)
            )

    # --- Apply (or skip in dry_run) ---
    if not dry_run:
        _apply_diff(wallet, diff, delete_missing=delete_missing)

    return diff


def _simulate_apply(wallet: "Wallet", diff: ImportDiff, *, delete_missing: bool):
    """Build a throwaway wallet-like object reflecting the post-import state.

    Used only for capital validation: we need compute_capital() to see
    the final state without actually mutating the real wallet.

    Returns a SimpleNamespace with .accounts, .entries and .sum(), which
    is the minimal contract compute_capital() requires.

    """
    from types import SimpleNamespace

    # Start from current wallet state, then apply the diff on local copies
    sim_accounts = list(wallet.accounts)
    sim_entries = list(wallet.entries)

    modified_account_mwids = {old.mwid for old, _ in diff.accounts_modify}
    sim_accounts = [a for a in sim_accounts if a.mwid not in modified_account_mwids]
    sim_accounts += [new for _, new in diff.accounts_modify]
    sim_accounts += diff.accounts_add
    if delete_missing:
        deleted = {a.mwid for a in diff.accounts_delete}
        sim_accounts = [a for a in sim_accounts if a.mwid not in deleted]

    modified_entry_mwids = {old.mwid for old, _ in diff.entries_modify}
    sim_entries = [e for e in sim_entries if e.mwid not in modified_entry_mwids]
    sim_entries += [new for _, new in diff.entries_modify]
    sim_entries += diff.entries_add
    if delete_missing:
        deleted = {e.mwid for e in diff.entries_delete}
        sim_entries = [e for e in sim_entries if e.mwid not in deleted]

    # Minimal sum() replacement walking the simulated entries
    def _sim_sum(account, date_range):
        total = Money(0)
        for e in sim_entries:
            if e.source == account:
                total = total + (e.amount * -1)
            elif e.target == account:
                total = total + (e.amount * +1)
        return total

    return SimpleNamespace(
        accounts=sim_accounts,
        entries=sim_entries,
        sum=_sim_sum,
    )


def _next_mwid(existing: list) -> int:
    """Return the next available MWID given a list of entities."""
    used = {e.mwid for e in existing if e.mwid >= 0}
    return max(used, default=-1) + 1


def _apply_diff(wallet: "Wallet", diff: ImportDiff, *, delete_missing: bool) -> None:
    """Apply the computed diff to `wallet` in place.

    New entities (mwid=-1) are assigned the next available MWID,
    mirroring what Wallet.write() does when persisting to disk. This
    ensures round-trips (export → import → export → import) don't
    duplicate entities in the absence of a write() call.

    """
    # Accounts
    modified_mwids = {old.mwid for old, _ in diff.accounts_modify}
    wallet.accounts = [a for a in wallet.accounts if a.mwid not in modified_mwids]
    wallet.accounts += [new for _, new in diff.accounts_modify]
    for new_a in diff.accounts_add:
        new_a.mwid = _next_mwid(wallet.accounts)
        wallet.accounts.append(new_a)
    if delete_missing:
        deleted = {a.mwid for a in diff.accounts_delete}
        wallet.accounts = [a for a in wallet.accounts if a.mwid not in deleted]

    # Categories
    modified_mwids = {old.mwid for old, _ in diff.categories_modify}
    wallet.categories = [c for c in wallet.categories if c.mwid not in modified_mwids]
    wallet.categories += [new for _, new in diff.categories_modify]
    for new_c in diff.categories_add:
        new_c.mwid = _next_mwid(wallet.categories)
        wallet.categories.append(new_c)
    if delete_missing:
        deleted = {c.mwid for c in diff.categories_delete}
        wallet.categories = [c for c in wallet.categories if c.mwid not in deleted]

    # Entries
    modified_mwids = {old.mwid for old, _ in diff.entries_modify}
    wallet.entries = [e for e in wallet.entries if e.mwid not in modified_mwids]
    wallet.entries += [new for _, new in diff.entries_modify]
    for new_e in diff.entries_add:
        new_e.mwid = _next_mwid(wallet.entries)
        wallet.entries.append(new_e)
    if delete_missing:
        deleted = {e.mwid for e in diff.entries_delete}
        wallet.entries = [e for e in wallet.entries if e.mwid not in deleted]

    # Counterparts: add any newly referenced ones
    existing_names = {cp.name for cp in wallet.counterparts}
    for cp in diff.counterparts_new:
        if cp.name not in existing_names:
            wallet.counterparts.append(cp)
            existing_names.add(cp.name)
