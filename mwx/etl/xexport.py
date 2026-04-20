# 2026/04/20
"""
export.py - Export wallet data to an Excel workbook.

Defines a single public function, `export()`, that writes a Wallet's
accounts, categories and entries to an .xlsx file with one sheet per
entity type, adapting each field to a native Excel representation
(dates as dates, amounts as numbers, colors as cell fills, booleans
as TRUE/FALSE, etc.).

The resulting workbook is designed to be editable and reimportable:
- MWID is the stable anchor; editing it breaks the link to the model.
- Derived fields (repr_name) are omitted to avoid ambiguity.
- Immutable fields (Type) are exported as read-only visual columns.
- All UI text is in Spanish for the end user; internal field names
  stay in English in code comments and parameters.

"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any, Callable, Sequence

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from mwx.etl.common import compute_capital

if TYPE_CHECKING:
    from mwx.wallet import Wallet


# Meta schema version. Bump if the meta sheet layout changes in a way
# that breaks backward compatibility with older import() implementations.
_META_SCHEMA = "mwx-export/1"
_META_SHEET_NAME = "__meta__"


# Styling constants

_FONT_NAME = "Calibri"
_HEADER_FILL = PatternFill("solid", start_color="FF305496")
_HEADER_FONT = Font(name=_FONT_NAME, bold=True, color="FFFFFFFF")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

_BODY_FONT = Font(name=_FONT_NAME)
_LEGACY_FONT = Font(name=_FONT_NAME, italic=True, color="FF808080")

_READONLY_FILL = PatternFill("solid", start_color="FFE8E8E8")
_MWID_FILL = PatternFill("solid", start_color="FFF2F2F2")

_DATE_FMT = "yyyy-mm-dd"
_AMOUNT_FMT = "#,##0.00"

_TOP_ALIGN = Alignment(vertical="top")

# Labels for type fields in the visible sheets
_TYPE_LABELS = {-1: "Gasto", 0: "Transferencia", +1: "Ingreso"}


# Column specifications
#
# Each spec: (header, extractor, cell_format, width, kind)
# kind ∈ {"normal", "mwid", "readonly", "color", "multiline"}
# - "normal":    plain value
# - "mwid":      MWID anchor; light grey fill, bold
# - "readonly":  immutable field; grey fill, comment "no editable"
# - "color":     value is a '#RRGGBB' string; paint cell bg and pick
#                contrasting font color automatically
# - "multiline": replace '\n' in the value with ' // ' for readability

_ColumnSpec = tuple[str, Callable[[Any], Any], str | None, int, str]


# Color utilities


def _hex_to_argb(hex_color: str) -> str:
    """Convert '#RRGGBB' to 'FFRRGGBB' (openpyxl expects ARGB)."""
    return "FF" + hex_color.lstrip("#").upper()


def _relative_luminance(hex_color: str) -> float:
    """Compute WCAG relative luminance of a '#RRGGBB' color, in [0, 1].

    Darker colors → lower luminance. Used to decide contrast font color.

    """
    h = hex_color.lstrip("#")
    r, g, b = (int(h[i : i + 2], 16) / 255.0 for i in (0, 2, 4))

    def _channel(c: float) -> float:
        return c / 12.92 if c <= 0.03928 else ((c + 0.055) / 1.055) ** 2.4

    return 0.2126 * _channel(r) + 0.7152 * _channel(g) + 0.0722 * _channel(b)


def _contrast_font(hex_color: str, base: Font) -> Font:
    """Return a Font like `base`, but white on dark backgrounds.

    Threshold 0.5 on WCAG relative luminance works well for typical
    palette colors; hex codes that don't parse fall back to `base`.

    """
    try:
        lum = _relative_luminance(hex_color)
    except (ValueError, IndexError):
        return base
    if lum < 0.5:
        return Font(
            name=base.name,
            bold=base.bold,
            italic=base.italic,
            color="FFFFFFFF",
        )
    return base


# Value utilities


def _flatten_newlines(value: Any) -> Any:
    """Replace newlines with ' // ' in strings; pass through otherwise."""
    if not isinstance(value, str):
        return value
    normalized = value.replace("\r\n", "\n").replace("\r", "\n")
    return normalized.replace("\n", " // ")


# Column specs per entity


def _account_specs() -> list[_ColumnSpec]:
    return [
        ("MWID", lambda a: a.mwid, None, 8, "mwid"),
        ("Nombre", lambda a: a.name, None, 24, "normal"),
        ("Orden", lambda a: a.order, None, 8, "normal"),
        ("Color", lambda a: a.color, None, 12, "color"),
        ("Visible", lambda a: a.is_visible, None, 9, "normal"),
        ("Legacy", lambda a: a.is_legacy, None, 9, "normal"),
    ]


def _category_specs() -> list[_ColumnSpec]:
    return [
        ("MWID", lambda c: c.mwid, None, 8, "mwid"),
        ("Código", lambda c: c.code, None, 8, "normal"),
        ("Nombre", lambda c: c.name, None, 28, "normal"),
        ("Tipo", lambda c: _TYPE_LABELS[c.type], None, 14, "readonly"),
        ("Icono", lambda c: c.icon_id, None, 8, "normal"),
        ("Color", lambda c: c.color, None, 12, "color"),
        ("Legacy", lambda c: c.is_legacy, None, 9, "normal"),
    ]


def _entry_specs() -> list[_ColumnSpec]:
    return [
        ("MWID", lambda e: e.mwid, None, 8, "mwid"),
        ("Fecha", lambda e: e.date, _DATE_FMT, 12, "normal"),
        ("Tipo", lambda e: _TYPE_LABELS[e.type], None, 14, "readonly"),
        ("Importe", lambda e: float(e.amount), _AMOUNT_FMT, 12, "normal"),
        ("Origen", lambda e: e.source.repr_name, None, 24, "normal"),
        ("Destino", lambda e: e.target.repr_name, None, 24, "normal"),
        ("Categoría", lambda e: e.category.repr_name, None, 32, "normal"),
        ("Concepto", lambda e: e.item, None, 30, "normal"),
        ("Detalles", lambda e: e.details, None, 45, "multiline"),
        ("Es factura", lambda e: e.is_bill, None, 10, "normal"),
    ]


# Sheet writer


def _write_sheet(
    ws: Worksheet,
    title: str,
    entities: Sequence[Any],
    specs: list[_ColumnSpec],
    *,
    legacy_attr: str | None = "is_legacy",
) -> None:
    """Populate a worksheet with entities according to column specs."""
    ws.title = title

    n_cols = len(specs)
    last_col_letter = get_column_letter(n_cols)

    # Row 1 - headers
    header_row = 1
    for col_idx, (header, _, _, width, _) in enumerate(specs, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 20

    # Data rows - sorted by entity's native sorting key
    sorted_entities = sorted(entities)
    first_data_row = header_row + 1

    for row_offset, entity in enumerate(sorted_entities):
        row_idx = first_data_row + row_offset
        is_legacy = bool(legacy_attr and getattr(entity, legacy_attr, False))
        row_font = _LEGACY_FONT if is_legacy else _BODY_FONT

        for col_idx, (header, extractor, fmt, _, kind) in enumerate(specs, start=1):
            try:
                value = extractor(entity)
            except Exception:  # noqa: BLE001 - never break export on one bad field
                value = None

            if kind == "multiline":
                value = _flatten_newlines(value)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = row_font
            cell.alignment = _TOP_ALIGN
            if fmt:
                cell.number_format = fmt

            # Kind-specific styling
            if kind == "mwid":
                cell.fill = _MWID_FILL
                cell.font = Font(
                    name=row_font.name,
                    bold=True,
                    italic=row_font.italic,
                    color=row_font.color,
                )
            elif kind == "readonly":
                cell.fill = _READONLY_FILL
            elif kind == "color" and isinstance(value, str) and value.startswith("#"):
                cell.fill = PatternFill("solid", start_color=_hex_to_argb(value))
                cell.font = _contrast_font(value, row_font)

    # Freeze header + MWID column; filter on headers
    ws.freeze_panes = "B2"
    if sorted_entities:
        last_row = first_data_row + len(sorted_entities) - 1
        ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{last_row}"

    # Comments on key header cells
    for col_idx, (_, _, _, _, kind) in enumerate(specs, start=1):
        if kind == "readonly":
            ws.cell(row=header_row, column=col_idx).comment = Comment(
                "Campo inmutable. No editar: al reimportar se ignorarán "
                "los cambios en esta columna.",
                "MWX",
            )
        elif kind == "mwid":
            ws.cell(row=header_row, column=col_idx).comment = Comment(
                "Identificador interno. No modificar filas existentes. "
                "Dejar en blanco para filas nuevas.",
                "MWX",
            )


# Meta sheet writer


def _mwx_version() -> str:
    """Best-effort read of mwx version; empty string if unavailable."""
    try:
        import mwx  # local import, may not be importable from tests

        return getattr(mwx, "__version__", "") or ""
    except Exception:  # noqa: BLE001
        return ""


def _write_meta_sheet(ws: Worksheet, wallet: "Wallet") -> None:
    """Write the hidden meta sheet with canonical validation values.

    Layout: two tables separated by a blank row.
    - Top table: generic metadata (schema, export date, version, etc.)
    - Bottom table: per-account capital, plus the grand total.

    The sheet is marked hidden so it doesn't appear by default.

    """
    ws.title = _META_SHEET_NAME

    total, per_account = compute_capital(wallet)

    # --- Top table: metadata ---
    meta_rows: list[tuple[str, Any]] = [
        ("schema", _META_SCHEMA),
        ("exported_at", datetime.now().isoformat(timespec="seconds")),
        ("mwx_version", _mwx_version()),
        ("source_path", str(wallet.source_path) if wallet.source_path else ""),
        ("total_capital", total),
        ("n_entries", len(wallet.entries)),
        ("n_accounts", len(wallet.accounts)),
        ("n_categories", len(wallet.categories)),
        ("n_counterparts", len(wallet.counterparts)),
    ]

    # Headers
    ws.cell(row=1, column=1, value="Meta").font = _HEADER_FONT
    ws.cell(row=1, column=1).fill = _HEADER_FILL
    ws.cell(row=1, column=2, value="Valor").font = _HEADER_FONT
    ws.cell(row=1, column=2).fill = _HEADER_FILL

    for offset, (key, val) in enumerate(meta_rows, start=2):
        ws.cell(row=offset, column=1, value=key).font = Font(name=_FONT_NAME, bold=True)
        cell = ws.cell(row=offset, column=2, value=val)
        cell.font = _BODY_FONT
        if key == "total_capital":
            cell.number_format = _AMOUNT_FMT

    # --- Separator + bottom table: per-account capital ---
    per_acc_header_row = len(meta_rows) + 3  # +1 (1-indexed) +1 blank row +1

    ws.cell(row=per_acc_header_row, column=1, value="Cuenta").font = _HEADER_FONT
    ws.cell(row=per_acc_header_row, column=1).fill = _HEADER_FILL
    ws.cell(row=per_acc_header_row, column=2, value="Capital").font = _HEADER_FONT
    ws.cell(row=per_acc_header_row, column=2).fill = _HEADER_FILL

    for offset, (repr_name, cap) in enumerate(per_account.items(), start=1):
        row = per_acc_header_row + offset
        ws.cell(row=row, column=1, value=repr_name).font = _BODY_FONT
        cell = ws.cell(row=row, column=2, value=cap)
        cell.font = _BODY_FONT
        cell.number_format = _AMOUNT_FMT

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 32

    # Hide the sheet
    ws.sheet_state = "hidden"


# Public API


def export(
    wallet: "Wallet",
    path: str | Path,
    *,
    overwrite: bool = False,
) -> Path:
    """Export wallet data to an Excel workbook.

    Creates a .xlsx file with three sheets: 'Movimientos', 'Categorías'
    and 'Cuentas'. Each sheet adapts its fields to native Excel types:
    dates as dates, amounts as numbers with thousand separators, colors
    painted as cell background (with automatic light/dark font contrast)
    and kept as hex string in the same cell, booleans as TRUE/FALSE,
    multiline strings flattened with ' // ' separators. Legacy rows are
    rendered in italic grey.

    The output is designed for editing and reimporting: MWID anchors
    each row, derived fields are omitted, and immutable fields (Type)
    are marked read-only with a visible grey background.

    Parameters
    ----------
    wallet:
        The Wallet instance to export.
    path:
        Destination path for the .xlsx file. If the extension is
        missing, '.xlsx' is appended.
    overwrite:
        If False (default) and the target exists, raises
        FileExistsError. If True, replaces the file.

    Returns the absolute path to the written file.

    """
    path = Path(path)
    if path.suffix.lower() != ".xlsx":
        path = path.with_suffix(".xlsx")
    if path.exists() and not overwrite:
        raise FileExistsError(
            f"Target file '{path}' already exists. Use overwrite=True to replace."
        )
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # Sheet 1 - Movimientos (active sheet, main focus)
    ws_entries = wb.active
    _write_sheet(ws_entries, "Movimientos", wallet.entries, _entry_specs())

    # Sheet 2 - Categorías
    ws_cats = wb.create_sheet("Categorías")
    _write_sheet(ws_cats, "Categorías", wallet.categories, _category_specs())

    # Sheet 3 - Cuentas
    ws_accs = wb.create_sheet("Cuentas")
    _write_sheet(ws_accs, "Cuentas", wallet.accounts, _account_specs())

    # Sheet 4 - __meta__ (hidden, for reimport validation)
    ws_meta = wb.create_sheet(_META_SHEET_NAME)
    _write_meta_sheet(ws_meta, wallet)

    wb.save(path)
    return path.resolve()
