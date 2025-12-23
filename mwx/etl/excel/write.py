# 2025/12/23
"""
write.py - Writing MyWallet data to Excel.

Defines function 'write', which writes MyWallet content to an Excel file, with
formatting and structure suitable for further processing.

It consists of three sheets: Entries, Accounts, Categories.

"""
from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING, Any

import openpyxl
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from mwx import Wallet


HEADER_ACCOUNTS = ["MWID", "Name", "Order", "Color", "Visible"]
HEADER_CATEGORIES = ["MWID", "Code", "Name", "Type", "Icon ID", "Color"]
HEADER_ENTRIES = [
    "MWID",
    "Year",
    "Month",
    "Day",
    "Amount",
    "Source",
    "Target",
    "Category",
    "Item",
    "Details",
]


XL_STYLE_NOTAS = "Note"
XL_ACC_COLUMNS_WIDTHS = [10, 22, 10, 14, 10]
XL_CAT_COLUMNS_WIDTHS = [10, 10, 27, 10, 10, 14]
XL_ENT_COLUMNS_WIDTHS = [10, 10, 10, 10, 15, 22, 22, 27, 25, 35]
XL_FORMAT_MONEY = '+_-* #,##0.00 €_-;-* #,##0.00 €_-;_-* "-"?? €_-;_-@_-'

CUSTOM_STYLE_HEADER = {
    "bold": True,
    "color": "CCCCCC",
    "align": "center",
}


def write(output_path: str | Path, data: Wallet, *, override: bool = False) -> None:
    """Writes MyWallet data to an Excel file.

    If `output_path` exists and `override` is False, raises an error.

    """
    # Check path
    output_path = Path(output_path)
    if output_path.exists() and not override:
        raise FileExistsError(f"File {output_path} already exists.")

    # Create workbook and sheets
    wb = openpyxl.Workbook()
    ws_entries = wb.active
    ws_entries.title = "Entries"
    ws_accounts = wb.create_sheet(title="Accounts")
    ws_categories = wb.create_sheet(title="Categories")

    # Accounts
    set_column_widths(ws_accounts, XL_ACC_COLUMNS_WIDTHS)
    write_and_style(
        ws_accounts,
        [1, 1],
        content=[
            {"value": header, **CUSTOM_STYLE_HEADER} for header in HEADER_ACCOUNTS
        ],
    )
    for i, account in enumerate(data.find(entity="account", is_legacy=False)):
        write_and_style(
            ws_accounts,
            [1, 2 + i],
            [
                {"value": account.mwid, "color": "E8E8E8"},
                {"value": account.name, "align": "left"},
                {"value": account.order},
                {"value": account.color},
                {"value": "SÍ" if account.is_visible else "NO"},
            ],
        )

    # Write Categories
    set_column_widths(ws_categories, XL_CAT_COLUMNS_WIDTHS)
    write_and_style(
        ws_categories,
        [1, 1],
        content=[
            {"value": header, **CUSTOM_STYLE_HEADER} for header in HEADER_CATEGORIES
        ],
    )
    for i, category in enumerate(data.find(entity="category", is_legacy=False)):
        _type = ["TRANS", "IN", "OUT"][category.type]
        write_and_style(
            ws_categories,
            [1, 2 + i],
            [
                {"value": category.mwid, "color": "E8E8E8"},
                {"value": category.code},
                {"value": category.name, "align": "left"},
                {"value": _type},
                {"value": category.icon_id},
                {"value": category.color},
            ],
        )

    # Write Entries
    set_column_widths(ws_entries, XL_ENT_COLUMNS_WIDTHS)
    write_and_style(
        ws_entries,
        [1, 1],
        content=[{"value": header, **CUSTOM_STYLE_HEADER} for header in HEADER_ENTRIES],
    )
    for i, entry in enumerate(data.find(entity="entry")):
        mwid = -entry.mwid if entry.type == 0 else entry.mwid
        write_and_style(
            ws_entries,
            [1, 2 + i],
            [
                {"value": mwid, "color": "E8E8E8"},
                {"value": entry.date.year},
                {"value": entry.date.month},
                {"value": entry.date.day},
                {"value": entry.amount.to_float(), "number_format": XL_FORMAT_MONEY},
                {"value": entry.source.repr_name, "align": "left"},
                {"value": entry.target.repr_name, "align": "left"},
                {"value": entry.category.repr_name, "align": "left"},
                {"value": entry.item, "align": "left"},
                {"value": entry.details.replace("\n", "$$"), "align": "left"},
            ],
        )

    # Save workbook
    wb.save(output_path)


# Auxiliar functions


def set_column_widths(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    widths: list[int],
) -> None:
    """Sets the column widths for the given sheet."""
    for i, width in enumerate(widths):
        sheet.column_dimensions[get_column_letter(i + 1)].width = width


def write_and_style(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    coords: list[int, int],
    content: list[Any],
) -> None:
    """Writes content to sheet and applies styling.

    Sets all cells to 'Notas' style, then fills with specified background color
    and writes the content.

    """
    x, y = coords
    for i, value in enumerate(content):
        # Value
        cell = sheet.cell(row=y, column=x + i)
        cell.value = value["value"]

        # Style
        cell.style = XL_STYLE_NOTAS
        cell.fill = openpyxl.styles.PatternFill(
            start_color=value.get("color", "FFFFFF"),
            end_color=value.get("color", "FFFFFF"),
            fill_type="solid",
        )
        if value.get("bold", False):
            cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(
            horizontal=value.get("align", "center"),
            vertical="center",
        )

        # Format
        if "number_format" in value:
            cell.number_format = value["number_format"]

        # Protection
        if value.get("protect", False):
            cell.protection = openpyxl.styles.Protection(locked=True)
        else:
            cell.protection = openpyxl.styles.Protection(locked=False)
